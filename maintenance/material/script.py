import os
import openpyxl
import csv


def ExcelToCsv():
    # 初始化一个空列表用于存储所有数据
    all_data = []

    # 获取当前脚本所在目录
    current_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(current_dir, 'material')
    print(f"path: {path}")

    # 获取所有Excel文件
    excel_files = [f for f in os.listdir(path) if f.endswith('.xlsx') or f.endswith('.xls')]
    print(f"excel_files: {excel_files}")
    # 遍历所有Excel文件
    for excel_file in excel_files:
        print(f"当前文件：{excel_file}")
        # 打开Excel文件
        workbook = openpyxl.load_workbook(os.path.join(path, excel_file))

        # 获取所有的工作表
        sheets = workbook.sheetnames

        # 遍历所有工作表
        for sheet in sheets:
            # 获取当前工作表
            worksheet = workbook[sheet]

            # 跳过标题行
            next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))

            # 遍历工作表的所有行
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                # 提取第1，4，5，6列的数据
                if row[0] is None:
                    continue
                data = [row[0], row[3], row[4], row[5]]
                print(f"\tdate:{data}")


                # 将数据添加到列表中
                all_data.append(data)


    # 将所有数据保存到CSV文件中
    with open(os.path.join(current_dir, 'output.csv'), 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)

        # 写入标题行
        csv_writer.writerow(['name', 'type', 'place', 'remark'])

        # 写入数据行
        for data in all_data:
            csv_writer.writerow(data)

    print("CSV文件已保存。")


if __name__ == "__main__":
    ExcelToCsv()
    print("Hello!")
